import random

import general
import requests
import lxml.html
import json
import openpyxl
import re
import csv
import pandas as pd
import datetime
from datetime import datetime
import os

headers={
'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'
}

# file_location="C:\\Users\\Aboli.Valvi\\PycharmProjects\\pythonProject\\ASOS_DE_SKU_ASSORTMENT_INPUT.xlsx"
# file_location="input_1.xlsx"
file_location="input_1.xlsx"


def html_cache_page_saving(sku, response):
    list_pagesave=[]
    datazone = datetime.now()
    f_date = datazone.strftime("%d_%m_%Y")
    strdate = datazone.day
    strm = datazone.month
    stry = datazone.year
    pageid = sku
    cpid = pageid + '_' + str(strdate) + '_' + str(strm) + '_' + str(stry)
    # ASS_folder = f"E:\ADIDAS_SavePages\Jdsports_uk\ASS"

    ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Asos_DE\\PDP"
    sos_date_wise_folder = ASS_folder + f"\\{f_date}"
    if os.path.exists(sos_date_wise_folder):
        pass
    else:
        os.mkdir(sos_date_wise_folder)
    sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
    sos_filename = sos_filename.replace("+", "_").replace("-", "_")
    page_path = sos_filename.replace('/', '')
    print(page_path)
    page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
                                  'https:////ecxus440.eclerx.com//cachepages//').replace('\\',
                                                                                         '//').replace(
    '//', '/')
    print(page_path)
    if os.path.exists(sos_filename):
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(response)
            f.close()
    else:
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(response)
            f.close()
    list_pagesave.append(page_path)
    # return page_path
    return list_pagesave

def get_xl_file(path):
    workbook=openpyxl.load_workbook(path)
    sheet_obj=workbook.active
    all_rows=sheet_obj.max_row

    urlsdata=[]
    for i in range(1, all_rows):
        # Adidas_id=sheet_obj.cell(row=i+1,column=2)
        # urls=sheet_obj.cell(row=i+1,column=1)
        # adidas_mpc = sheet_obj.cell(row=i + 1, column=3)
        # u_data=str(urls.value)+'&*&'+str(Adidas_id.value)+'&*&'+str(adidas_mpc.value)
        # urlsdata.append(u_data.split('&*&'))

        pdp_url = sheet_obj.cell(row=i + 1, column=4)
        sku_id = sheet_obj.cell(row=i + 1, column=3)
        country = sheet_obj.cell(row=i + 1, column=2)
        website_name = sheet_obj.cell(row=i + 1, column=1)

        u_data = [pdp_url.value, sku_id.value, country.value, website_name.value]
        urlsdata.append(u_data)
    print(urlsdata)
    return urlsdata
get_xl_file(path=file_location)

from csv import DictWriter
import general
now = datetime.now()
# date_time = now.strftime("%Y-%m-%dT%H:%M:%SZ")
date_time1 = str(now.strftime("%m%d%Y"))
def data_exp(dict):
    fieldnames = ['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID',
                                                   'Product_URL',
                                                   'Product_Name', 'Category_Path', 'Specification', 'Description',
                                                   'Currency', 'List_Price',
                                                   'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count',
                                                   'Image_URLs', 'Variant',
                                                   'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name',
                                                   'Stock_Count', 'Stock_Condition',
                                                   'Stock_Message', 'Sustainability_Badge', 'Reason_Code',
                                                   'Crawling_TimeStamp', 'Cache_Page_Link',
                                                   'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5']
    with open(f'ASOS_DE1_sku_{date_time1}_rehit2.csv', 'a+', encoding='UTF-8-sig', newline='') as file:
    # with open('ASOS_DE1_sku_0330_pending2.csv', 'a+', encoding='UTF-8-sig', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
        # writer.writeheader()
        writer.writerows(dict)
        file.close()

def listToString(s):
    # initialize an empty string
    str1 = " "

    # return string
    return (str1.join(s))


def join_string1(list_string):
    # Join the string based on '-' delimiter
    string = '|'.join(list_string)

    return string

listofdata=[]
import pandas as pd
products_df1=pd.DataFrame()
def get_data(n):
    global products_df1
    p1 = ['154.28.67.106','154.28.67.111','154.28.67.116','154.28.67.117','154.28.67.125','154.28.67.131','154.28.67.133','154.28.67.142','154.28.67.156','154.28.67.163','154.28.67.173','154.28.67.18','154.28.67.182','154.28.67.184','154.28.67.20','154.28.67.200','154.28.67.210','154.28.67.218','154.28.67.222','154.28.67.223','154.28.67.231','154.28.67.240','154.28.67.243','154.28.67.253','154.28.67.39','154.28.67.4','154.28.67.49','154.28.67.5','154.28.67.61','154.28.67.80','154.28.67.81','154.28.67.87','154.28.67.88','154.28.67.96','154.28.67.99','154.7.230.100','154.7.230.101','154.7.230.103','154.7.230.107','154.7.230.109','154.7.230.130','154.7.230.132','154.7.230.14','154.7.230.140','154.7.230.147','154.7.230.151','154.7.230.156','154.7.230.163','154.7.230.170','154.7.230.18','154.7.230.183','154.7.230.188','154.7.230.189','154.7.230.19','154.7.230.190','154.7.230.198','154.7.230.204','154.7.230.209','154.7.230.235','154.7.230.238','154.7.230.246','154.7.230.29','154.7.230.41','154.7.230.42','154.7.230.51','154.7.230.55','154.7.230.60','154.7.230.61','154.7.230.74','154.7.230.82','154.7.230.89','23.131.8.112','23.131.8.115','23.131.8.117','23.131.8.12','23.131.8.121','23.131.8.124','23.131.8.150','23.131.8.161','23.131.8.166','23.131.8.171','23.131.8.173','23.131.8.176','23.131.8.177','23.131.8.181','23.131.8.19','23.131.8.192','23.131.8.194','23.131.8.199','23.131.8.202','23.131.8.203','23.131.8.204','23.131.8.207','23.131.8.209','23.131.8.213','23.131.8.216','23.131.8.225','23.131.8.228','23.131.8.231','23.131.8.238','23.131.8.254','23.131.8.36','23.131.8.5','23.131.8.76','23.131.8.93','23.131.8.95','23.131.8.99','23.131.88.105','23.131.88.12','23.131.88.137','23.131.88.139','23.131.88.140','23.131.88.145','23.131.88.150','23.131.88.151','23.131.88.153','23.131.88.154','23.131.88.156','23.131.88.165','23.131.88.18','23.131.88.191','23.131.88.192','23.131.88.194','23.131.88.198','23.131.88.202','23.131.88.206','23.131.88.220','23.131.88.223','23.131.88.228','23.131.88.233','23.131.88.24','23.131.88.242','23.131.88.244','23.131.88.47','23.131.88.63','23.131.88.67','23.131.88.73','23.131.88.80','23.131.88.81','23.131.88.82','23.131.88.88','23.131.88.97','23.170.144.149','23.170.144.209','23.170.144.212','23.170.144.242','23.170.144.83','23.170.145.117','23.170.145.167','23.170.145.182','23.170.145.19','23.170.145.203','23.226.17.101','23.226.17.109','23.226.17.112','23.226.17.113','23.226.17.115','23.226.17.123','23.226.17.129','23.226.17.143','23.226.17.148','23.226.17.165','23.226.17.186','23.226.17.199','23.226.17.201','23.226.17.207','23.226.17.210','23.226.17.219','23.226.17.220','23.226.17.222','23.226.17.229','23.226.17.250','23.226.17.254','23.226.17.26','23.226.17.33','23.226.17.4','23.226.17.49','23.226.17.5','23.226.17.55','23.226.17.66','23.226.17.7','23.226.17.72','23.226.17.78','23.226.17.8','23.226.17.86','23.226.17.90','23.226.17.93','23.230.177.105','23.230.177.110','23.230.177.113','23.230.177.121','23.230.177.130','23.230.177.14','23.230.177.143','23.230.177.15','23.230.177.150','23.230.177.154','23.230.177.165','23.230.177.173','23.230.177.191','23.230.177.196','23.230.177.203','23.230.177.206','23.230.177.208','23.230.177.217','23.230.177.220','23.230.177.221','23.230.177.224','23.230.177.228','23.230.177.231','23.230.177.235','23.230.177.237','23.230.177.241','23.230.177.27','23.230.177.38','23.230.177.52','23.230.177.61','23.230.177.67','23.230.177.72','23.230.177.80','23.230.177.88','23.230.177.94','23.230.177.99','23.230.197.103','23.230.197.106','23.230.197.109','23.230.197.11','23.230.197.12','23.230.197.122','23.230.197.124','23.230.197.146','23.230.197.155','23.230.197.156','23.230.197.174','23.230.197.179','23.230.197.181','23.230.197.196','23.230.197.2','23.230.197.201','23.230.197.207','23.230.197.208','23.230.197.225','23.230.197.227','23.230.197.233','23.230.197.236','23.230.197.239','23.230.197.240','23.230.197.244','23.230.197.251','23.230.197.50','23.230.197.52','23.230.197.54','23.230.197.60','23.230.197.71','23.230.197.80','23.230.197.81','23.230.197.84','23.230.197.97','23.230.74.102','23.230.74.110','23.230.74.116','23.230.74.125','23.230.74.133','23.230.74.135','23.230.74.14','23.230.74.141','23.230.74.149','23.230.74.15','23.230.74.157','23.230.74.16','23.230.74.170','23.230.74.172','23.230.74.174','23.230.74.183','23.230.74.187','23.230.74.19','23.230.74.198','23.230.74.208','23.230.74.212','23.230.74.215','23.230.74.23','23.230.74.230','23.230.74.231','23.230.74.252','23.230.74.30','23.230.74.41','23.230.74.57','23.230.74.58','23.230.74.59','23.230.74.6','23.230.74.75','23.230.74.81','23.230.74.88','23.230.74.91','23.27.222.108','23.27.222.109','23.27.222.134','23.27.222.138','23.27.222.159','23.27.222.161','23.27.222.164','23.27.222.166','23.27.222.178','23.27.222.19','23.27.222.195','23.27.222.201','23.27.222.202','23.27.222.203','23.27.222.208','23.27.222.21','23.27.222.211','23.27.222.218','23.27.222.223','23.27.222.228','23.27.222.234','23.27.222.236','23.27.222.242','23.27.222.251','23.27.222.253','23.27.222.34','23.27.222.61','23.27.222.62','23.27.222.69','23.27.222.70','23.27.222.72','23.27.222.73','23.27.222.74','23.27.222.81','23.27.222.93','38.131.131.110','38.131.131.114','38.131.131.123','38.131.131.125','38.131.131.137','38.131.131.142','38.131.131.145','38.131.131.147','38.131.131.15','38.131.131.154','38.131.131.16','38.131.131.17','38.131.131.173','38.131.131.18','38.131.131.193','38.131.131.204','38.131.131.207','38.131.131.227','38.131.131.229','38.131.131.233','38.131.131.238','38.131.131.246','38.131.131.248','38.131.131.250','38.131.131.31','38.131.131.36','38.131.131.50','38.131.131.58','38.131.131.64','38.131.131.70','38.131.131.71','38.131.131.74','38.131.131.83','38.131.131.94','38.131.131.99','38.75.75.104','38.75.75.111','38.75.75.112','38.75.75.119','38.75.75.120','38.75.75.123','38.75.75.127','38.75.75.139','38.75.75.14','38.75.75.143','38.75.75.155','38.75.75.156','38.75.75.158','38.75.75.170','38.75.75.179','38.75.75.188','38.75.75.2','38.75.75.201','38.75.75.231','38.75.75.232','38.75.75.241','38.75.75.246','38.75.75.251','38.75.75.26','38.75.75.29','38.75.75.4','38.75.75.44','38.75.75.49','38.75.75.56','38.75.75.58','38.75.75.62','38.75.75.72','38.75.75.76','38.75.75.79','38.75.75.88','38.96.156.108','38.96.156.112','38.96.156.128','38.96.156.131','38.96.156.14','38.96.156.142','38.96.156.143','38.96.156.149','38.96.156.16','38.96.156.163','38.96.156.165','38.96.156.169','38.96.156.186','38.96.156.188','38.96.156.190','38.96.156.192','38.96.156.194','38.96.156.199','38.96.156.218','38.96.156.236','38.96.156.240','38.96.156.252','38.96.156.28','38.96.156.32','38.96.156.35','38.96.156.56','38.96.156.57','38.96.156.6','38.96.156.67','38.96.156.77','38.96.156.80','38.96.156.83','38.96.156.84','38.96.156.89','38.96.156.92','45.238.157.100','45.238.157.104','45.238.157.106','45.238.157.110','45.238.157.116','45.238.157.118','45.238.157.119','45.238.157.12','45.238.157.123','45.238.157.132','45.238.157.14','45.238.157.149','45.238.157.15','45.238.157.183','45.238.157.186','45.238.157.189','45.238.157.2','45.238.157.212','45.238.157.214','45.238.157.217','45.238.157.22','45.238.157.228','45.238.157.23','45.238.157.247','45.238.157.43','45.238.157.48','45.238.157.51','45.238.157.52','45.238.157.53','45.238.157.56','45.238.157.61','45.238.157.65','45.238.157.72','45.238.157.79','45.238.157.8','45.238.159.103','45.238.159.107','45.238.159.110','45.238.159.114','45.238.159.116','45.238.159.123','45.238.159.126','45.238.159.144','45.238.159.148','45.238.159.15','45.238.159.156','45.238.159.165','45.238.159.167','45.238.159.183','45.238.159.20','45.238.159.208','45.238.159.217','45.238.159.220','45.238.159.23','45.238.159.230','45.238.159.235','45.238.159.237','45.238.159.238','45.238.159.24','45.238.159.249','45.238.159.251','45.238.159.32','45.238.159.34','45.238.159.51','45.238.159.6','45.238.159.66','45.238.159.77','45.238.159.79','45.238.159.82','45.238.159.91']
    global response
    for i in get_xl_file(path=file_location):
        #print(i)
        INPUT_url = i[0]
        print(INPUT_url)
        SKUID = i[1]
        # if 'Asos_DE_prd' in SKUID or 'Asos_DE_grp' in SKUID:
        #     SKUID= str(INPUT_url).split('?clr')[0].split('/')[-1]
        print(INPUT_url,' ------ ',SKUID)
        try:
            p_auth = str("csimonra:h19VA2xZ")
            p_host = random.choice(p1)
            p_port = "29842"
            proxy = {
                'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
            }
            response = requests.get(INPUT_url, headers=headers,proxies=proxy,timeout=30)
            tree = lxml.html.fromstring(response.text)

            cache_data = response.text

            html_file_path = html_cache_page_saving(sku=str(i[1]), response=cache_data)


            if response.status_code == 200:
                reason_Code = "SucessPF"
                print('status code', response.status_code)
            # else:
            #     raise Exception('Failed to fetch web page' + INPUT_url)


                p_data = ''.join(re.findall(r'window.asos.pdp.config.product = (.*?);', response.text))
                if not p_data:
                    print(INPUT_url, 'not found')


                try:
                    product_json = json.loads(p_data)
                    #print(product_json)
                    # try:
                    #     sell_name = str(product_json['products'][0]['variants']['seller'])
                    # except Exception as e:
                    #     sell_name='-'
                    try:
                        a = tree.xpath('//script[contains(text(),"window.asos.pdp.config.product = ")]/text()')
                        # print(type(a))
                        d = str(a)
                        # print(d)
                        data1 = d.split("window.asos.pdp.config.product =")[1].split(";")[0]
                        # print(data1)
                        data1=str(data1).replace("\\'","'")
                        j_data = json.loads(data1)
                        print(type(j_data))
                        # varxy = j_data['products']
                        var = j_data['products'][0]['variants']
                        try:
                            sell_name = var[0].get('seller')
                            if not sell_name:
                                sell_name='-'
                        except Exception as e:
                            sell_name = "-"
                    except Exception as e:
                        sell_name='-'
                    Product_name = product_json['name']
                    # print(Product_name)

                    product_brand = product_json['brandName']
                    # print(product_brand)

                    now1 = datetime.now()
                    output_date = datetime.now().strftime("%m-%d-%YT%H:%M:%SZ")
                    print(output_date)

                    cat = tree.xpath('//*[@aria-label="breadcrumbs"]//li//text()')
                    if not cat:
                        breadcrum = str(response.text).split("breadcrumb: '")[-1].split("version: '")[0].strip().replace("',","")
                        Category_Path = breadcrum.replace('/', '>').replace('>>','>').strip()
                    else:
                        Category_Path = "".join([i.strip() for i in cat if i.strip()]).strip()
                        print(Category_Path)

                    image = []
                    img = len(product_json['images'])
                    for im in range(0, img):
                        imgs = product_json['images'][im]['url']
                        image.append(imgs)
                    Images = '|'.join(image)
                    # print(Images)
                    try:
                        color_of_variant = product_json['variants'][0]['colour']
                        # print(color_of_variant)
                    except:
                        try:
                            color_of_variant=product_json['products'][0]['variants'][0]['colour']
                        except:
                            color_of_variant ='-'
                    variant = []
                    var = len(product_json['variants'])
                    for v in range(0, var):
                        vars = product_json['variants'][v]['size']
                        variant.append(vars)
                    variants = '|'.join(variant)
                    varientSize=variants.split('|')
                    # print(variant)

                    variant = []
                    stockvar = []
                    variant_id=[]
                    seller_name=[]
                    var = len(product_json['variants'])
                    for v in range(0, var):
                        vars = product_json['variants'][v]['size']
                        stock = product_json['variants'][v]['isInStock']
                        vid = product_json['variants'][v]['variantId']
                        sellname = product_json['variants'][v]['seller']
                        stockvarc = 'instock' if stock == True else 'out of stock'

                        if stockvarc == "out of stock":
                            stockvarc = "Out of Stock"

                        if stockvarc == "instock":
                            stockvarc = "InStock"

                        stockvar.append(stockvarc)
                        variant.append(vars)
                        variant_id.append(vid)
                        seller_name.append(sellname)
                    variants = '|'.join(variant)
                    varientSize = variants.split('|')
                    # print(variant)
                    # print(stockvar)


                    Description = '|'.join(tree.xpath('//*[@class="product-description"]//ul/li//text()'))
                    # print(Description)


                    ####        specification
                    # Specifications = '|'.join(tree.xpath('//*[@class="about-me"]/p/text()'))
                    #
                    # spe = Specifications if Specifications else '|'.join(tree.xpath('//*[@class="about-me"]/text()[preceding-sibling::br]'))

                    specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/text()')
                    xyz = listToString(specifications)
                    if ":" in xyz:
                        for abc in specifications:
                            if ':' in abc:
                                specifications = abc.strip(' ').strip('\n').strip('.')
                    else:
                        specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/p/text()')
                        xyz = listToString(specifications)
                        if ":" in xyz:
                            for abc in specifications:
                                if ':' in abc:
                                    specifications = abc.strip(' ').strip('\n').strip('.')
                        else:
                            specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/div/text()')
                            xyz = listToString(specifications)
                            if ":" in xyz:
                                for abc in specifications:
                                    if ":" in abc:
                                        specifications = abc.strip(' ').strip('\n').strip('.')
                            else:
                                specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/p/span/text()')
                                xyz = listToString(specifications)
                                if ':' in xyz:
                                    for abc in specifications:
                                        if ":" in abc:
                                            specifications = abc.strip(' ').strip('\n').strip('.')
                                else:
                                    specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/div/div/text()')
                                    xyz = listToString(specifications)
                                    if ':' in xyz:
                                        for abc in specifications:
                                            if ':' in abc:
                                                specifications = abc.strip(' ').strip('\n').strip('.')
                                    else:
                                        specifications = tree.xpath('//div[@class="col"]/div[@class="about-me"]/span/text()')
                                        xyz = listToString(specifications)
                                        if ':' in xyz:
                                            for abc in specifications:
                                                if ':' in abc:
                                                    specifications = abc.strip(' ').strip('\n').strip('.')
                                        else:
                                            specifications = tree.xpath(
                                                '//div[@class="col"]/div[@class="about-me"]/div/span/text()')
                                            xyz = listToString(specifications)
                                            if ':' in xyz:
                                                for abc in specifications:
                                                    if ':' in abc:
                                                        specifications = abc.strip(' ').strip('\n').strip('.')

                    if len(specifications) == 0:
                        specifications = '-'

                    if specifications != '-':
                        specifications_value = specifications.split('| ')
                        specifications_bool = specifications_value[-1][0].isdigit()
                        if specifications_bool == True:
                            specifications_value[-2] = specifications_value[-2] + ", " + specifications_value[-1]
                            specifications_value.pop()
                            specifications = specifications_value
                            specifications = join_string1(specifications)

                    try:
                        Product_id = product_json['productCode']
                    except:
                        Product_id=INPUT_url.split('?')[0].split('/')[-1]


                    # product_url = tree.xpath('//*[@rel="canonical"]/@href')[0]

                    try:
                        riview_data = ''.join(re.findall(r'window.asos.pdp.config.ratings = (.*?);', response.text))
                        review_json = json.loads(riview_data)
                    except Exception as e:
                        review_json=''

                    try:
                        Ratings = review_json['averageOverallRating']
                    except:
                        Ratings='-'
                    try:
                        Review = review_json['totalReviewCount']
                    except:
                        Review='-'
                    # print(Ratings, Review)

                    try:
                        product_id_specific = product_json['images'][0]['productId']
                        # print(product_id_specific)
                        price_link = f'https://www.asos.com/api/product/catalogue/v3/stockprice?productIds={product_id_specific}&store=ES&currency=EUR&keyStoreDataversion=hgk0y12-29'

                        for i in range(1, 5):
                            try:
                                p_auth = str("csimonra:h19VA2xZ")
                                p_host = random.choice(p1)
                                p_port = "29842"
                                proxy = {
                                    'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                                    'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                                }
                                response2 = requests.request("GET", price_link, headers=headers, proxies=proxy, timeout=30)
                                if response2.status_code == 200:
                                    data = response2.text
                                    price_json = json.loads(data)
                                    # print(price_json)
                                    List_price = price_json[0]['productPrice']['previous']['text']
                                    Promo_price = price_json[0]['productPrice']['current']['text']

                                    a = List_price.replace('€', '')

                                    b = Promo_price.replace('€', '')
                                    # print(a, b)
                                    print('list and promo price done')
                                    break
                                else:
                                    a = '-'
                                    b = '-'

                            except Exception as e:
                                a = '-'
                                b = '-'

                        disurlfr = f'https://www.asos.com/api/product/catalogue/v3/stockprice?productIds={product_id_specific}&store=DE&currency=EUR&keyStoreDataversion=hgk0y12-29'

                        for i in range(1, 5):
                            try:
                                p_auth = str("csimonra:h19VA2xZ")
                                p_host = random.choice(p1)
                                p_port = "29842"
                                proxy = {
                                    'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
                                    'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
                                }
                                response1 = requests.request("GET", disurlfr, headers=headers, proxies=proxy, timeout=30)
                                if response1.status_code == 200:
                                    data = response1.text
                                    dis_json = json.loads(data)
                                    # print(dis_json)
                                    z = dis_json[0]['discountPercentage']
                                    dis = (str(z) + '%')
                                    print('list and promo price done.....')
                                    break
                                else:
                                    dis = '-'
                            except Exception as e:
                                dis = '-'
                    except Exception as e:
                        a,b,dis='-','-','-'
                        print(e)
                except Exception as e:
                    continue

                if variant !=[]:
                    for item, st, varid,sname in zip(variant, stockvar,variant_id,seller_name):
                        product_info = {
                            'SKU_ID':SKUID,
                            # 'INPUT_PLATFORMID': product_id_specific,
                            # 'INPUT_ADIDASID': 'Not Available',
                            'Website': 'ASOS',
                            'Country': 'DE',
                            'RPC':Product_id,
                            'MPC':'Not Available',
                            'Product_ID': Product_id,
                            # 'Product_URL': product_url,
                            'Product_URL': INPUT_url,
                            'Product_Name': general.clean(Product_name),
                            'Category_Path': general.clean(Category_Path),
                            'Specification': general.clean(specifications),
                            'Description': general.clean(Description),
                            'Currency':'EURO',
                            'List_Price':  a,
                            'Promo_Price': b,
                            'Discount': dis,
                            'Brand': general.clean(product_brand),
                            'Rating_Count': Ratings,
                            'Review_Count': Review,
                            'Image_URLs': Images,
                            'Variant': general.clean(item),
                            'Variant_ID': varid,
                            'Colour_of_Variant': general.clean(color_of_variant),
                            'Colour_Grouping': 'Not Available',
                            'Seller_Name': general.clean(sname),
                            'Stock_Count': '-',
                            'Stock_Condition': st,
                            'Stock_Message':'Not Available',
                            # 'Shipping_price': Shipping_price,
                            'Sustainability_Badge': 'Not Available',
                            'Reason_Code':reason_Code,
                            'Crawling_TimeStamp': output_date,
                            'Cache_Page_Link':  html_file_path[0],
                            'Extra1': '-',
                            'Extra2': '-',
                            'Extra3': '-',
                            'Extra4': '-',
                            'Extra5': '-',

                        }
                        print(product_info)
                        data_exp(dict=[product_info])
                    # listofdata.append(product_info)
                    # products_df1 = products_df1.append(product_info, ignore_index=True)
                    # writer = pd.ExcelWriter(f'ASOS_DE_PDP1.xlsx', engine='xlsxwriter',
                    #                         options={'strings_to_urls': False})
                    # products_df1.to_excel(writer,
                    #                       columns=['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID',
                    #                                'Product_URL',
                    #                                'Product_Name', 'Category_Path', 'Specification', 'Description',
                    #                                'Currency', 'List_Price',
                    #                                'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count',
                    #                                'Image_URLs', 'Variant',
                    #                                'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name',
                    #                                'Stock_Count', 'Stock_Condition',
                    #                                'Stock_Message', 'Sustainability_Badge', 'Reason_Code',
                    #                                'Crawling_TimeStamp', 'Cache_Page_Link',
                    #                                'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'], index=False)
                    # writer.save()
                else:
                    product_info = {
                        'SKU_ID': SKUID,
                        # 'INPUT_PLATFORMID': product_id_specific,
                        # 'INPUT_ADIDASID': 'Not Available',
                        'Website': 'ASOS',
                        'Country': 'DE',
                        'RPC': Product_id,
                        'MPC': 'Not Available',
                        'Product_ID': Product_id,
                        # 'Product_URL': product_url,
                        'Product_URL': INPUT_url,
                        'Product_Name': general.clean(Product_name),
                        'Category_Path': general.clean(Category_Path),
                        'Specification': general.clean(specifications),
                        'Description': general.clean(Description),
                        'Currency': 'EURO',
                        'List_Price': a,
                        'Promo_Price': b,
                        'Discount': dis,
                        'Brand': general.clean(product_brand),
                        'Rating_Count': Ratings,
                        'Review_Count': Review,
                        'Image_URLs': Images,
                        'Variant': '-',
                        'Variant_ID': '-',
                        'Colour_of_Variant': general.clean(color_of_variant),
                        'Colour_Grouping': 'Not Available',
                        'Seller_Name': general.clean(sell_name),
                        'Stock_Count': '-',
                        'Stock_Condition': 'In Stock',
                        'Stock_Message': 'Not Available',
                        # 'Shipping_price': Shipping_price,
                        'Sustainability_Badge': 'Not Available',
                        'Reason_Code': reason_Code,
                        'Crawling_TimeStamp': output_date,
                        'Cache_Page_Link': html_file_path[0],
                        'Extra1': '-',
                        'Extra2': '-',
                        'Extra3': '-',
                        'Extra4': '-',
                        'Extra5': '-',

                    }
                    print(product_info)
                    data_exp(dict=[product_info])
                # print(listofdata)

                # to_csv=listofdata
                # keys=to_csv[0].keys()
                # with open('ASOS_DE_PDP1.csv','w',encoding='utf-8-sig',newline="")as output_file:
                # # with open('ASOS_DE_PDP1.xlsx','w',encoding='utf-8',newline="")as output_file:
                #     dict_writer=csv.DictWriter(output_file,keys)
                #     dict_writer.writeheader()
                #     dict_writer.writerows(to_csv)

            # # except Exception as e:
            # #     print('EXCEPTOPM ', e)

            elif n != 0:
                n = n - 1
                get_data(n)
                import time
                time.sleep(0.5)

            elif response.status_code == 408:
                output_date = datetime.now().strftime("%m-%d-%YT%H:%M:%SZ")
                t1 = str(output_date).replace('-13-', '-12-')
                t2 = str(output_date).replace('-13-', '-11-')
                print(output_date)
                reason_Code = 'TimeOut'
                product_info = {
                    'SKU_ID': SKUID,
                    # 'INPUT_PLATFORMID': '-',
                    # 'INPUT_ADIDASID': '-',
                    'Website': 'ASOS',
                    'Country': 'DE',
                    'RPC': '-',
                    'MPC': 'Not Available',
                    'Product_ID': '-',
                    'Product_URL': INPUT_url,
                    'Product_Name': '-',
                    'Category_Path': '-',
                    'Specification': '-',
                    'Description': '-',
                    'Currency': '-',
                    'List_Price': '-',
                    'Promo_Price': '-',
                    'Discount': '-',
                    'Brand': '-',
                    'Rating_Count': '-',
                    'Review_Count': '-',
                    'Image_URLs': '-',
                    'Variant': '-',
                    'Variant_ID': '-',
                    'Colour_of_Variant': '-',
                    'Colour_Grouping': '-',
                    'Seller_Name': '-',
                    'Stock_Count': '-',
                    'Stock_Condition': '-',
                    'Stock_Message': '-',
                    # 'Shipping_price': Shipping_price,
                    'Sustainability_Badge': '-',
                    'Reason_Code': reason_Code,
                    'Crawling_TimeStamp': output_date,
                    'Cache_Page_Link': html_file_path[0],
                    'Extra1': '-',
                    'Extra2': '-',
                    'Extra3': '-',
                    'Extra4': '-',
                    'Extra5': '',

                }
                data_exp(dict=[product_info])
                # listofdata.append(product_info)
                # products_df1 = products_df1.append(product_info, ignore_index=True)
                # writer = pd.ExcelWriter(f'ASOS_DE_PDP1.xlsx', engine='xlsxwriter',
                #                         options={'strings_to_urls': False})
                # products_df1.to_excel(writer,
                #                       columns=['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID',
                #                                'Product_URL',
                #                                'Product_Name', 'Category_Path', 'Specification', 'Description',
                #                                'Currency', 'List_Price',
                #                                'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count',
                #                                'Image_URLs', 'Variant',
                #                                'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name',
                #                                'Stock_Count', 'Stock_Condition',
                #                                'Stock_Message', 'Sustainability_Badge', 'Reason_Code',
                #                                'Crawling_TimeStamp', 'Cache_Page_Link',
                #                                'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'], index=False)
                # writer.save()
                # print(listofdata)
                #
                # to_csv = listofdata
                # keys = to_csv[0].keys()
                # with open('ASOS_DE_PDP1.csv', 'w', encoding='utf-8-sig', newline="") as output_file:
                # # with open('ASOS_DE_PDP1.xlsx', 'w', encoding='utf-8', newline="") as output_file:
                #     dict_writer = csv.DictWriter(output_file, keys)
                #     dict_writer.writeheader()
                #     dict_writer.writerows(to_csv)
            elif response.status_code == 403:
                output_date = datetime.now().strftime("%m-%d-%YT%H:%M:%SZ")
                t1 = str(output_date).replace('-13-', '-12-')
                t2 = str(output_date).replace('-13-', '-11-')
                reason_Code = 'Blocked'
                product_info = {
                    'SKU_ID': SKUID,
                    # 'INPUT_PLATFORMID': '-',
                    # 'INPUT_ADIDASID': '-',
                    'Website': 'ASOS',
                    'Country': 'DE',
                    'RPC': '-',
                    'MPC': 'Not Available',
                    'Product_ID': '-',
                    'Product_URL': INPUT_url,
                    'Product_Name': '-',
                    'Category_Path': '-',
                    'Specification': '-',
                    'Description': '-',
                    'Currency': '-',
                    'List_Price': '-',
                    'Promo_Price': '-',
                    'Discount': '-',
                    'Brand': '-',
                    'Rating_Count': '-',
                    'Review_Count': '-',
                    'Image_URLs': '-',
                    'Variant': '-',
                    'Variant_ID': '-',
                    'Colour_of_Variant': '-',
                    'Colour_Grouping': '-',
                    'Seller_Name': '-',
                    'Stock_Count': '-',
                    'Stock_Condition': '-',
                    'Stock_Message': '-',
                    # 'Shipping_price': Shipping_price,
                    'Sustainability_Badge': '-',
                    'Reason_Code': reason_Code,
                    'Crawling_TimeStamp':output_date,
                    'Cache_Page_Link': html_file_path[0],
                    'Extra1': '-',
                    'Extra2': '-',
                    'Extra3': '-',
                    'Extra4': '-',
                    'Extra5': '-',

                }
                data_exp(dict=[product_info])
                # listofdata.append(product_info)
                # products_df1 = products_df1.append(product_info, ignore_index=True)
                # writer = pd.ExcelWriter(f'ASOS_DE_PDP1.xlsx', engine='xlsxwriter',
                #                         options={'strings_to_urls': False})
                # products_df1.to_excel(writer,
                #                       columns=['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID',
                #                                'Product_URL',
                #                                'Product_Name', 'Category_Path', 'Specification', 'Description',
                #                                'Currency', 'List_Price',
                #                                'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count',
                #                                'Image_URLs', 'Variant',
                #                                'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name',
                #                                'Stock_Count', 'Stock_Condition',
                #                                'Stock_Message', 'Sustainability_Badge', 'Reason_Code',
                #                                'Crawling_TimeStamp', 'Cache_Page_Link',
                #                                'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'], index=False)
                # writer.save()
                # print(listofdata)
                #
                # to_csv = listofdata
                # keys = to_csv[0].keys()
                # with open('ASOS_DE_PDP1.csv', 'w', encoding='utf-8-sig',newline="") as output_file:
                # # with open('ASOS_DE_PDP1.xlsx', 'w', encoding='utf-8',newline="") as output_file:
                #     dict_writer = csv.DictWriter(output_file, keys)
                #     dict_writer.writeheader()
                #     dict_writer.writerows(to_csv)
            else:
                output_date = datetime.now().strftime("%m-%d-%YT%H:%M:%SZ")
                t1 = str(output_date).replace('-13-', '-12-')
                t2 = str(output_date).replace('-13-', '-11-')
                reason_Code = 'Sucess PNF'
                product_info = {
                    'SKU_ID': SKUID,
                    # 'INPUT_PLATFORMID': '-',
                    # 'INPUT_ADIDASID': '-',
                    'Website': 'ASOS',
                    'Country': 'DE',
                    'RPC': '-',
                    'MPC': 'Not Available',
                    'Product_ID': '-',
                    'Product_URL': INPUT_url,
                    'Product_Name': '-',
                    'Category_Path': '-',
                    'Specification': '-',
                    'Description': '-',
                    'Currency': '-',
                    'List_Price': '-',
                    'Promo_Price': '-',
                    'Discount': '-',
                    'Brand': '-',
                    'Rating_Count': '-',
                    'Review_Count': '-',
                    'Image_URLs': '-',
                    'Variant': '-',
                    'Variant_ID': '-',
                    'Colour_of_Variant': '-',
                    'Colour_Grouping': '-',
                    'Seller_Name': '-',
                    'Stock_Count': '-',
                    'Stock_Condition': '-',
                    'Stock_Message': '-',
                    # 'Shipping_price': Shipping_price,
                    'Sustainability_Badge': '-',
                    'Reason_Code': reason_Code,
                    'Crawling_TimeStamp': output_date,
                    'Cache_Page_Link': html_file_path[0],
                    'Extra1': '-',
                    'Extra2': '-',
                    'Extra3': '-',
                    'Extra4': '-',
                    'Extra5': '-',

                }
                data_exp(dict=[product_info])
                # listofdata.append(product_info)
                # products_df1 = products_df1.append(product_info, ignore_index=True)
                # writer = pd.ExcelWriter(f'ASOS_DE_PDP1.xlsx', engine='xlsxwriter',
                #                         options={'strings_to_urls': False})
                # products_df1.to_excel(writer,
                #                       columns=['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID',
                #                                'Product_URL',
                #                                'Product_Name', 'Category_Path', 'Specification', 'Description',
                #                                'Currency', 'List_Price',
                #                                'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count',
                #                                'Image_URLs', 'Variant',
                #                                'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name',
                #                                'Stock_Count', 'Stock_Condition',
                #                                'Stock_Message', 'Sustainability_Badge', 'Reason_Code',
                #                                'Crawling_TimeStamp', 'Cache_Page_Link',
                #                                'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'], index=False)
                # writer.save()
                # print(listofdata)
                #
                # to_csv = listofdata
                # keys = to_csv[0].keys()
                # with open('ASOS_DE_PDP1.csv', 'w', encoding='utf-8-sig',newline="") as output_file:
                # # with open('ASOS_DE_PDP1.xlsx', 'w', encoding='utf-8',newline="") as output_file:
                #     dict_writer = csv.DictWriter(output_file, keys)
                #     dict_writer.writeheader()
                #     dict_writer.writerows(to_csv)
        except Exception as e:
            output_date = datetime.now().strftime("%m-%d-%YT%H:%M:%SZ")
            t1 = str(output_date).replace('-13-', '-12-')
            t2 = str(output_date).replace('-13-', '-11-')
            print('timeout')
            reason_Code = 'timeout'
            product_info = {
                'SKU_ID': SKUID,
                # 'INPUT_PLATFORMID': '-',
                # 'INPUT_ADIDASID': '-',
                'Website': 'ASOS',
                'Country': 'DE',
                'RPC': '-',
                'MPC': 'Not Available',
                'Product_ID': '-',
                'Product_URL': INPUT_url,
                'Product_Name': '-',
                'Category_Path': '-',
                'Specification': '-',
                'Description': '-',
                'Currency': '-',
                'List_Price': '-',
                'Promo_Price': '-',
                'Discount': '-',
                'Brand': '-',
                'Rating_Count': '-',
                'Review_Count': '-',
                'Image_URLs': '-',
                'Variant': '-',
                'Variant_ID': '-',
                'Colour_of_Variant': '-',
                'Colour_Grouping': '-',
                'Seller_Name': '-',
                'Stock_Count': '-',
                'Stock_Condition': '-',
                'Stock_Message': '-',
                # 'Shipping_price': Shipping_price,
                'Sustainability_Badge': '-',
                'Reason_Code': reason_Code,
                'Crawling_TimeStamp': output_date,
                'Cache_Page_Link': '-',
                'Extra1': '-',
                'Extra2': '-',
                'Extra3': '-',
                'Extra4': '-',
                'Extra5': '',

            }
            data_exp(dict=[product_info])
            # listofdata.append(product_info)
            # products_df1 = products_df1.append(product_info, ignore_index=True)
            # writer = pd.ExcelWriter(f'ASOS_DE_PDP1.xlsx', engine='xlsxwriter',
            #                         options={'strings_to_urls': False})
            # products_df1.to_excel(writer,
            #                       columns=['SKU_ID', 'Website', 'Country', 'RPC', 'MPC', 'Product_ID',
            #                                'Product_URL',
            #                                'Product_Name', 'Category_Path', 'Specification', 'Description',
            #                                'Currency', 'List_Price',
            #                                'Promo_Price', 'Discount', 'Brand', 'Rating_Count', 'Review_Count',
            #                                'Image_URLs', 'Variant',
            #                                'Variant_ID', 'Colour_of_Variant', 'Colour_Grouping', 'Seller_Name',
            #                                'Stock_Count', 'Stock_Condition',
            #                                'Stock_Message', 'Sustainability_Badge', 'Reason_Code',
            #                                'Crawling_TimeStamp', 'Cache_Page_Link',
            #                                'Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'], index=False)
            # writer.save()
            # print(listofdata)
            #
            # to_csv = listofdata
            # keys = to_csv[0].keys()
            # with open('ASOS_DE_PDP1.csv', 'w',encoding='utf-8-sig', newline="") as output_file:
            # # with open('ASOS_DE_PDP1.xlsx', 'w',encoding='utf-8', newline="") as output_file:
            #     dict_writer = csv.DictWriter(output_file, keys)
            #     dict_writer.writeheader()
            #     dict_writer.writerows(to_csv)


get_data(n=3)
