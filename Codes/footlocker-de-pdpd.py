import random
import re

import general
from lxml import html
import time
import os
from csv import DictWriter
import re
import openpyxl
import requests
import lxml

import time
import datetime
# from adidas_proxy import AdidasProxy
from datetime import datetime
import json
import re
import uuid
import lxml
from lxml import html
from csv import DictWriter
import openpyxl
import time
from datetime import datetime
import json
import requests
import datetime
import datetime
start = time.time()

import requests
from lxml import html
import time
import json
from datetime import datetime, timedelta
import datetime
import openpyxl
# from adidas_proxy import AdidasProxy

start=time.time()
DominNam = 'https://www.footlocker.de'

headers={'authority': 'www.footlocker.de',
'method': 'GET',
'path': '/',
'scheme': 'https',
'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
'accept-encoding': 'gzip, deflate, br',
'accept-language': 'de;q=0.9',
'cache-control': 'max-age=0',
'cookie': 'at_check=true; AMCVS_40A3741F578E26BA7F000101%40AdobeOrg=1; AMCV_40A3741F578E26BA7F000101%40AdobeOrg=-1124106680%7CMCIDTS%7C19017%7CMCMID%7C25219398567362809234446184149299936138%7CMCAAMLH-1643605434%7C6%7CMCAAMB-1643605434%7CRKhpRz8krg2tLO6pguXWp5olkAcUniQYPHaMWWgdJ3xzPWQmdj0y%7CMCOPTOUT-1643007834s%7CNONE%7CvVersion%7C5.2.0; se=aks; s_pr_tbe65=1643000646767; s_cc=true; JSESSIONID=knr1c9c420dsn1so9wy8l70m.fzcxwefapipdb028882; _ga=GA1.3.782676145.1643000655; _gid=GA1.3.677682276.1643000655; userStatus=guest; userVIP=unknown; _crbx=84083e89-bc93-4a2c-ba7b-1a826b49103e; sc.ASP.NET_SESSIONID=kfljgsvjpar34jnqikfanr5n; sc.UserId=60d78174-425d-4483-b727-a19de0127119; s_sq=%5B%5BB%5D%5D; OptanonAlertBoxClosed=2022-01-24T05:05:33.283Z; __zlcmid=18CkeauVQ6CjXJG; s_pr_tbe66=1643002330541; fita.sid.footlocker_eu=mQxLQLzp_9xxO0IJ8g51MXBwEW5qBELA; datadome=zwir22~g-9ee6osivlyBvu4IzGwntWv4tiY6_X4uFMYmOKdkjBE~-r0zEHI~Ujd2YQXfBNRnyqIVOJWwxAo7uZacB5kj.lBHOp1BMzz3jwAo1P1FybBFzNdf39Y38iv; aa_pageHistory=[{"n":"","t":"","p":""},{"n":"uk: FLUK: W: Homepage","t":"Home","p":"/"}]; s_vs=1; s_lv=1643006451104; s_lv_s=Less%20than%201%20day; mbox=PC#33fb68f5a51f40da80d8b864e9442d49.38_0#1706251252|session#759905ec7a274eee95809696641e1888#1643008311; OptanonConsent=isGpcEnabled=0&datestamp=Mon+Jan+24+2022+12%3A10%3A52+GMT%2B0530+(India+Standard+Time)&version=6.23.0&isIABGlobal=false&hosts=&genVendors=&landingPath=NotLandingPage&groups=C0001%3A1%2CC0002%3A1%2CC0003%3A1%2CC0004%3A0&geolocation=%3B&AwaitingReconsent=false',
'if-none-match': 'W/"85f7f-b5UfyZa/SmcWjgNX6QOyuzZwRO0"',
'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
'sec-ch-ua-mobile': '?0',
'sec-ch-ua-platform': "Windows",
'sec-fetch-dest': 'document',
'sec-fetch-mode': 'navigate',
'sec-fetch-site': 'none',
'sec-fetch-user': '?1',
'upgrade-insecure-requests': '1',
'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36'}



# file_path='C:\\Users\\Piyush.Raut\\PycharmProjects\\pythonProject\\INPUTS\\Footlocker DE\\PDP Input Template.xlsx'
# file_path='ASSORT_input.xlsx'
file_path='ASSORT_input.xlsx'
def extract_xl():
    workbook= openpyxl.load_workbook(file_path)
    sheet_obj= workbook.active
    all_rows=sheet_obj.max_row

    inputdata=[]
    for i in range(1,all_rows):
        limit=sheet_obj.cell(row=i+1,column=3)
        urls= sheet_obj.cell(row=i+1,column=4)

        u_data=[urls.value,limit.value]

        inputdata.append(u_data)
    print(inputdata)
    return inputdata
import pandas as pd
products_df1=pd.DataFrame()
from csv import DictWriter
def store_data_by_product_id(dict):
    global products_df1
    fieldnames =['SKU_ID',	'Website',	'Country',	'RPC',	'MPC',	'Product_ID',	'Product_URL',
                 'Product_Name',	'Category_Path',	'Specification',	'Description',	'Currency',
                 'List_Price',	'Promo_Price',	'Discount',	'Brand',	'Rating_Count',	'Review_Count',
                 'Image_URLs',	'Variant',	'Variant_ID',	'Colour_of_Variant',	'Colour_Grouping',
                 'Seller_Name',	'Stock_Count',	'Stock_Condition',	'Stock_Message',	'Sustainability_Badge',
                 'Reason_Code',	'Crawling_TimeStamp',	'Cache_Page_Link',	'Extra1',	'Extra2',	'Extra3',
                 'Extra4',	'Extra5']

   # fieldnames = ['INPUT_PLATFORMID', 'INPUT_ADIDASID', 'Website Name', 'Country','Product_ID','Product_URL','Product_Name','Category_Path','Specification','Description','Currency','List_Price','Promo_Price','Discount','Brand','Rating_Count','Review_Count','Image_URLs','Variant','Variant_ID','Colour_of_Variant','Colour_Grouping','Seller_Name','Stock_Count','Stock_Condition','Stock_Message','Sustainability_Badge','Reason_Code','Crawling_TimeStamp','Cache_Page_Link','Extra1','Extra2','Extra3','Extra4','Extra5']
    with open('Footlocker_de1_2204.csv', 'a+', encoding='utf-8-sig', newline='') as file:
        writer = DictWriter(file, fieldnames=fieldnames)
       # writer.writeheader()
        writer.writerows(dict)
        file.close()

    # products_df1 = products_df1.append(dict, ignore_index=True)
    # writer = pd.ExcelWriter(f'footlocker_de_PDP_19032022_rehit1.xlsx', engine='xlsxwriter',
    #                         options={'strings_to_urls': False})
    # products_df1.to_excel(writer,
    #                       columns=['SKU_ID',	'Website',	'Country',	'RPC',	'MPC',	'Product_ID',	'Product_URL',
    #              'Product_Name',	'Category_Path',	'Specification',	'Description',	'Currency',
    #              'List_Price',	'Promo_Price',	'Discount',	'Brand',	'Rating_Count',	'Review_Count',
    #              'Image_URLs',	'Variant',	'Variant_ID',	'Colour_of_Variant',	'Colour_Grouping',
    #              'Seller_Name',	'Stock_Count',	'Stock_Condition',	'Stock_Message',	'Sustainability_Badge',
    #              'Reason_Code',	'Crawling_TimeStamp',	'Cache_Page_Link',	'Extra1',	'Extra2',	'Extra3',
    #              'Extra4',	'Extra5'], index=False)
    # writer.save()


def html_cache_page_saving(sku,saving_data):
    # page_path_list=[]
    datazone = datetime.datetime.now()
    f_date = datazone.strftime("%d_%m_%Y")
    strdate = datazone.day
    strm = datazone.month
    stry = datazone.year
    pageid = sku
    cpid = pageid + '_' + str(strdate) + '_' + str(strm) + '_' + str(stry)
    # ASS_folder = f"E:\ADIDAS_SavePages\Jdsports_uk\ASS"

    ASS_folder = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Footlocker_DE\\PDP"
    sos_date_wise_folder = ASS_folder + f"\\{f_date}"
    if os.path.exists(sos_date_wise_folder):
        pass
    else:
        os.mkdir(sos_date_wise_folder)
    sos_filename = sos_date_wise_folder + "\\" + cpid + ".html"
    sos_filename = sos_filename.replace("+", "_").replace("-", "_")
    page_path = sos_filename.replace('/', '')
    print(page_path)
    page_path = page_path.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\','https:////ecxus440.eclerx.com//cachepages//').replace('\\','//').replace('//', '/')
    print(page_path)
    if os.path.exists(sos_filename):
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(saving_data)
    else:
        with open(sos_filename, 'w', encoding='utf-8') as f:
            f.write(saving_data)
    # page_path_list.append(page_path)
    #
    # #pagesave12
    # cpid12 = pageid + '_12_03_2022'# + str(strdate) + '_' + str(strm) + '_' + str(stry)
    # ASS_folder12 = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Footlocker_DE\\PDP"
    # f_date12='12_03_2022'
    # sos_date_wise_folder12 = ASS_folder12 + f"\\{f_date12}"
    # if os.path.exists(sos_date_wise_folder12):
    #     pass
    # else:
    #     os.mkdir(sos_date_wise_folder12)
    # sos_filename12 = sos_date_wise_folder12 + "\\" + cpid12 + ".html"
    # sos_filename12 = sos_filename12.replace("+", "_").replace("-", "_")
    # page_path12 = sos_filename12.replace('/', '')
    # print(page_path12)
    # page_path12 = page_path12.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
    #                               'https:////ecxus440.eclerx.com//cachepages//').replace('\\', '//').replace('//', '/')
    # print(page_path12)
    # if os.path.exists(sos_filename12):
    #     with open(sos_filename12, 'w', encoding='utf-8') as f:
    #         f.write(saving_data)
    # else:
    #     with open(sos_filename12, 'w', encoding='utf-8') as f:
    #         f.write(saving_data)
    # page_path_list.append(page_path12)
    #
    # # pagesave11
    # cpid11 = pageid + '_11_03_2022'
    # ASS_folder11 = f"\\\\ecxus440\\E$\\ADIDAS_SavePages\\Footlocker_DE\\PDP"
    # f_date11 = '11_03_2022'
    # sos_date_wise_folder11 = ASS_folder11 + f"\\{f_date11}"
    # if os.path.exists(sos_date_wise_folder11):
    #     pass
    # else:
    #     os.mkdir(sos_date_wise_folder11)
    # sos_filename11 = sos_date_wise_folder11 + "\\" + cpid11 + ".html"
    # sos_filename11 = sos_filename11.replace("+", "_").replace("-", "_")
    # page_path11 = sos_filename11.replace('/', '')
    # print(page_path11)
    # page_path11 = page_path11.replace('\\\\ecxus440\\E$\\ADIDAS_SavePages\\',
    #                                   'https:////ecxus440.eclerx.com//cachepages//').replace('\\', '//').replace(
    #     '//', '/')
    # print(page_path11)
    # if os.path.exists(sos_filename11):
    #     with open(sos_filename11, 'w', encoding='utf-8') as f:
    #         f.write(saving_data)
    # else:
    #     with open(sos_filename11, 'w', encoding='utf-8') as f:
    #         f.write(saving_data)
    # page_path_list.append(page_path11)
    return page_path

# lacoste-powercourt-2-0-menshoes
#count=0
# pr = AdidasProxy()
#now = datetime
dom = 'https://www.footlocker.de'
def by_product_id(prod_id_data,n):
    global details_dict
    count = 0
    #  = 'https://www.footlocker.co.uk/en/product/~/{}.html'.format(prod_id_data[0])
    # new=
    # product_id_url= #'https://www.footlocker.de/de/product/lacoste-l001-menshoes/{}.html'.format(str(prod_id_data[0]))
    # # product_id_url='https://en.aboutyou.de/p/about-you-x-kevin-trapp/between-seasons-coat-armin-7424647?is_s=none&is_h=sl'
    # print(product_id_url)
   # print(prod_id_data[0])
    product_url = prod_id_data[0]
    print('rrrrrrrrrrrrrr',product_url)
    count=count+1
    print(count)
    #exit()
    #pro
    proxy_iplum = ["lum-customer-c_127755f5-zone-us_zone-ip-205.237.95.118",
                   "lum-customer-c_127755f5-zone-us_zone-ip-205.237.94.12",
                   "lum-customer-c_127755f5-zone-us_zone-ip-205.237.93.15",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.138.24",
                   "lum-customer-c_127755f5-zone-us_zone-ip-66.56.81.81",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.101.195",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.185.214",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.166.124",
                   "lum-customer-c_127755f5-zone-us_zone-ip-216.19.221.179",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.174.51",
                   "lum-customer-c_127755f5-zone-us_zone-ip-74.85.208.20",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.114.58",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.119.236",
                   "lum-customer-c_127755f5-zone-us_zone-ip-198.240.89.44",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.199.211",
                   "lum-customer-c_127755f5-zone-us_zone-ip-199.244.60.208",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.120.29",
                   "lum-customer-c_127755f5-zone-us_zone-ip-74.85.210.146",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.124.31",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.115.29",
                   "lum-customer-c_127755f5-zone-us_zone-ip-198.240.101.9",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.126.198",
                   "lum-customer-c_127755f5-zone-us_zone-ip-67.213.122.166",
                   "lum-customer-c_127755f5-zone-us_zone-ip-216.19.200.134",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.240.143",
                   "lum-customer-c_127755f5-zone-us_zone-ip-216.19.199.1",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.13.196.121",
                   "lum-customer-c_127755f5-zone-us_zone-ip-91.92.218.14",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.113.199",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.131.157",
                   "lum-customer-c_127755f5-zone-us_zone-ip-46.232.209.111",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.13.217.58",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.13.193.211",
                   "lum-customer-c_127755f5-zone-us_zone-ip-180.149.17.222",
                   "lum-customer-c_127755f5-zone-us_zone-ip-203.78.175.112",
                   "lum-customer-c_127755f5-zone-us_zone-ip-188.119.117.166",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.12.184.31",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.59.134",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.13.200.246",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.12.203.25",
                   "lum-customer-c_127755f5-zone-us_zone-ip-180.149.25.80",
                   "lum-customer-c_127755f5-zone-us_zone-ip-188.211.24.139",
                   "lum-customer-c_127755f5-zone-us_zone-ip-91.192.215.74",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.12.202.151",
                   "lum-customer-c_127755f5-zone-us_zone-ip-180.149.6.40",
                   "lum-customer-c_127755f5-zone-us_zone-ip-180.149.26.196",
                   "lum-customer-c_127755f5-zone-us_zone-ip-91.92.217.112",
                   "lum-customer-c_127755f5-zone-us_zone-ip-193.200.104.140",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.54.25",
                   "lum-customer-c_127755f5-zone-us_zone-ip-180.149.2.89",
                   "lum-customer-c_127755f5-zone-us_zone-ip-206.204.38.62",
                   "lum-customer-c_127755f5-zone-us_zone-ip-213.188.83.143",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.53.18",
                   "lum-customer-c_127755f5-zone-us_zone-ip-91.245.235.235",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.85.212",
                   "lum-customer-c_127755f5-zone-us_zone-ip-78.138.40.246",
                   "lum-customer-c_127755f5-zone-us_zone-ip-185.246.173.58",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.12.198.227",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.57.142",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.60.18",
                   "lum-customer-c_127755f5-zone-us_zone-ip-89.38.132.172",
                   "lum-customer-c_127755f5-zone-us_zone-ip-213.188.76.146",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.12.182.89",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.51.212",
                   "lum-customer-c_127755f5-zone-us_zone-ip-208.86.196.158",
                   "lum-customer-c_127755f5-zone-us_zone-ip-168.151.179.15",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.12.188.46",
                   "lum-customer-c_127755f5-zone-us_zone-ip-213.188.75.89",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.12.190.164",
                   "lum-customer-c_127755f5-zone-us_zone-ip-213.188.68.39",
                   "lum-customer-c_127755f5-zone-us_zone-ip-89.40.81.29",
                   "lum-customer-c_127755f5-zone-us_zone-ip-185.223.56.108",
                   "lum-customer-c_127755f5-zone-us_zone-ip-161.129.160.90",
                   "lum-customer-c_127755f5-zone-us_zone-ip-213.188.88.130",
                   "lum-customer-c_127755f5-zone-us_zone-ip-152.39.153.185",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.13.205.38",
                   "lum-customer-c_127755f5-zone-us_zone-ip-203.78.174.235",
                   "lum-customer-c_127755f5-zone-us_zone-ip-206.204.49.106",
                   "lum-customer-c_127755f5-zone-us_zone-ip-94.176.61.225",
                   "lum-customer-c_127755f5-zone-us_zone-ip-180.149.23.229",
                   "lum-customer-c_127755f5-zone-us_zone-ip-95.215.38.210",
                   "lum-customer-c_127755f5-zone-us_zone-ip-209.95.161.115",
                   "lum-customer-c_127755f5-zone-us_zone-ip-162.43.236.118",
                   "lum-customer-c_127755f5-zone-us_zone-ip-45.142.97.199",
                   "lum-customer-c_127755f5-zone-us_zone-ip-152.39.160.7",
                   "lum-customer-c_127755f5-zone-us_zone-ip-119.13.221.75",
                   "lum-customer-c_127755f5-zone-us_zone-ip-139.5.105.86",
                   "lum-customer-c_127755f5-zone-us_zone-ip-180.149.0.122",
                   "lum-customer-c_127755f5-zone-us_zone-ip-206.204.5.224",
                   "lum-customer-c_127755f5-zone-us_zone-ip-203.109.62.114",
                   "lum-customer-c_127755f5-zone-us_zone-ip-139.5.107.122",
                   "lum-customer-c_127755f5-zone-us_zone-ip-84.39.228.157",
                   "lum-customer-c_127755f5-zone-us_zone-ip-203.109.63.54",
                   "lum-customer-c_127755f5-zone-us_zone-ip-95.215.37.169",
                   "lum-customer-c_127755f5-zone-us_zone-ip-162.43.235.85",
                   "lum-customer-c_127755f5-zone-us_zone-ip-152.39.214.9",
                   "lum-customer-c_127755f5-zone-us_zone-ip-110.238.215.203",
                   "lum-customer-c_127755f5-zone-us_zone-ip-162.43.229.39",
                   "lum-customer-c_127755f5-zone-us_zone-ip-185.10.4.83",
                   "lum-customer-c_127755f5-zone-us_zone-ip-216.194.92.180"]

    try:



        port = '22225'
        rand_ips = "zproxy.lum-superproxy.io"
        rndusername = random.choice(proxy_iplum)
        usern_passw = rndusername + ':dngrv4oofa9a'
        proxy = {'https': "https://{}@{}:{}/".format(usern_passw, rand_ips, port)}
        p1 = ['154.28.67.106', '154.28.67.111', '154.28.67.116', '154.28.67.117', '154.28.67.125',
              '154.28.67.131',
              '154.28.67.133', '154.28.67.142', '154.28.67.156', '154.28.67.163', '154.28.67.173',
              '154.28.67.18',
              '154.28.67.182', '154.28.67.184', '154.28.67.20', '154.28.67.200', '154.28.67.210',
              '154.28.67.218',
              '154.28.67.222', '154.28.67.223', '154.28.67.231', '154.28.67.240', '154.28.67.243',
              '154.28.67.253',
              '154.28.67.39', '154.28.67.4', '154.28.67.49', '154.28.67.5', '154.28.67.61', '154.28.67.80',
              '154.28.67.81', '154.28.67.87', '154.28.67.88', '154.28.67.96', '154.28.67.99', '154.7.230.100',
              '154.7.230.101', '154.7.230.103', '154.7.230.107', '154.7.230.109', '154.7.230.130',
              '154.7.230.132',
              '154.7.230.14', '154.7.230.140', '154.7.230.147', '154.7.230.151', '154.7.230.156',
              '154.7.230.163',
              '154.7.230.170', '154.7.230.18', '154.7.230.183', '154.7.230.188', '154.7.230.189',
              '154.7.230.19',
              '154.7.230.190', '154.7.230.198', '154.7.230.204', '154.7.230.209', '154.7.230.235',
              '154.7.230.238',
              '154.7.230.246', '154.7.230.29', '154.7.230.41', '154.7.230.42', '154.7.230.51', '154.7.230.55',
              '154.7.230.60', '154.7.230.61', '154.7.230.74', '154.7.230.82', '154.7.230.89', '23.131.8.112',
              '23.131.8.115', '23.131.8.117', '23.131.8.12', '23.131.8.121', '23.131.8.124', '23.131.8.150',
              '23.131.8.161', '23.131.8.166', '23.131.8.171', '23.131.8.173', '23.131.8.176', '23.131.8.177',
              '23.131.8.181', '23.131.8.19', '23.131.8.192', '23.131.8.194', '23.131.8.199', '23.131.8.202',
              '23.131.8.203', '23.131.8.204', '23.131.8.207', '23.131.8.209', '23.131.8.213', '23.131.8.216',
              '23.131.8.225', '23.131.8.228', '23.131.8.231', '23.131.8.238', '23.131.8.254', '23.131.8.36',
              '23.131.8.5', '23.131.8.76', '23.131.8.93', '23.131.8.95', '23.131.8.99', '23.131.88.105',
              '23.131.88.12', '23.131.88.137', '23.131.88.139', '23.131.88.140', '23.131.88.145',
              '23.131.88.150',
              '23.131.88.151', '23.131.88.153', '23.131.88.154', '23.131.88.156', '23.131.88.165',
              '23.131.88.18',
              '23.131.88.191', '23.131.88.192', '23.131.88.194', '23.131.88.198', '23.131.88.202',
              '23.131.88.206',
              '23.131.88.220', '23.131.88.223', '23.131.88.228', '23.131.88.233', '23.131.88.24',
              '23.131.88.242',
              '23.131.88.244', '23.131.88.47', '23.131.88.63', '23.131.88.67', '23.131.88.73', '23.131.88.80',
              '23.131.88.81', '23.131.88.82', '23.131.88.88', '23.131.88.97', '23.170.144.149',
              '23.170.144.209',
              '23.170.144.212', '23.170.144.242', '23.170.144.83', '23.170.145.117', '23.170.145.167',
              '23.170.145.182', '23.170.145.19', '23.170.145.203', '23.226.17.101', '23.226.17.109',
              '23.226.17.112', '23.226.17.113', '23.226.17.115', '23.226.17.123', '23.226.17.129',
              '23.226.17.143',
              '23.226.17.148', '23.226.17.165', '23.226.17.186', '23.226.17.199', '23.226.17.201',
              '23.226.17.207',
              '23.226.17.210', '23.226.17.219', '23.226.17.220', '23.226.17.222', '23.226.17.229',
              '23.226.17.250',
              '23.226.17.254', '23.226.17.26', '23.226.17.33', '23.226.17.4', '23.226.17.49', '23.226.17.5',
              '23.226.17.55', '23.226.17.66', '23.226.17.7', '23.226.17.72', '23.226.17.78', '23.226.17.8',
              '23.226.17.86', '23.226.17.90', '23.226.17.93', '23.230.177.105', '23.230.177.110',
              '23.230.177.113',
              '23.230.177.121', '23.230.177.130', '23.230.177.14', '23.230.177.143', '23.230.177.15',
              '23.230.177.150', '23.230.177.154', '23.230.177.165', '23.230.177.173', '23.230.177.191',
              '23.230.177.196', '23.230.177.203', '23.230.177.206', '23.230.177.208', '23.230.177.217',
              '23.230.177.220', '23.230.177.221', '23.230.177.224', '23.230.177.228', '23.230.177.231',
              '23.230.177.235', '23.230.177.237', '23.230.177.241', '23.230.177.27', '23.230.177.38',
              '23.230.177.52', '23.230.177.61', '23.230.177.67', '23.230.177.72', '23.230.177.80',
              '23.230.177.88',
              '23.230.177.94', '23.230.177.99', '23.230.197.103', '23.230.197.106', '23.230.197.109',
              '23.230.197.11', '23.230.197.12', '23.230.197.122', '23.230.197.124', '23.230.197.146',
              '23.230.197.155', '23.230.197.156', '23.230.197.174', '23.230.197.179', '23.230.197.181',
              '23.230.197.196', '23.230.197.2', '23.230.197.201', '23.230.197.207', '23.230.197.208',
              '23.230.197.225', '23.230.197.227', '23.230.197.233', '23.230.197.236', '23.230.197.239',
              '23.230.197.240', '23.230.197.244', '23.230.197.251', '23.230.197.50', '23.230.197.52',
              '23.230.197.54', '23.230.197.60', '23.230.197.71', '23.230.197.80', '23.230.197.81',
              '23.230.197.84',
              '23.230.197.97', '23.230.74.102', '23.230.74.110', '23.230.74.116', '23.230.74.125',
              '23.230.74.133',
              '23.230.74.135', '23.230.74.14', '23.230.74.141', '23.230.74.149', '23.230.74.15',
              '23.230.74.157',
              '23.230.74.16', '23.230.74.170', '23.230.74.172', '23.230.74.174', '23.230.74.183',
              '23.230.74.187',
              '23.230.74.19', '23.230.74.198', '23.230.74.208', '23.230.74.212', '23.230.74.215',
              '23.230.74.23',
              '23.230.74.230', '23.230.74.231', '23.230.74.252', '23.230.74.30', '23.230.74.41', '23.230.74.57',
              '23.230.74.58', '23.230.74.59', '23.230.74.6', '23.230.74.75', '23.230.74.81', '23.230.74.88',
              '23.230.74.91', '23.27.222.108', '23.27.222.109', '23.27.222.134', '23.27.222.138',
              '23.27.222.159',
              '23.27.222.161', '23.27.222.164', '23.27.222.166', '23.27.222.178', '23.27.222.19',
              '23.27.222.195',
              '23.27.222.201', '23.27.222.202', '23.27.222.203', '23.27.222.208', '23.27.222.21',
              '23.27.222.211',
              '23.27.222.218', '23.27.222.223', '23.27.222.228', '23.27.222.234', '23.27.222.236',
              '23.27.222.242',
              '23.27.222.251', '23.27.222.253', '23.27.222.34', '23.27.222.61', '23.27.222.62', '23.27.222.69',
              '23.27.222.70', '23.27.222.72', '23.27.222.73', '23.27.222.74', '23.27.222.81', '23.27.222.93',
              '38.131.131.110', '38.131.131.114', '38.131.131.123', '38.131.131.125', '38.131.131.137',
              '38.131.131.142', '38.131.131.145', '38.131.131.147', '38.131.131.15', '38.131.131.154',
              '38.131.131.16', '38.131.131.17', '38.131.131.173', '38.131.131.18', '38.131.131.193',
              '38.131.131.204', '38.131.131.207', '38.131.131.227', '38.131.131.229', '38.131.131.233',
              '38.131.131.238', '38.131.131.246', '38.131.131.248', '38.131.131.250', '38.131.131.31',
              '38.131.131.36', '38.131.131.50', '38.131.131.58', '38.131.131.64', '38.131.131.70',
              '38.131.131.71',
              '38.131.131.74', '38.131.131.83', '38.131.131.94', '38.131.131.99', '38.75.75.104',
              '38.75.75.111',
              '38.75.75.112', '38.75.75.119', '38.75.75.120', '38.75.75.123', '38.75.75.127', '38.75.75.139',
              '38.75.75.14', '38.75.75.143', '38.75.75.155', '38.75.75.156', '38.75.75.158', '38.75.75.170',
              '38.75.75.179', '38.75.75.188', '38.75.75.2', '38.75.75.201', '38.75.75.231', '38.75.75.232',
              '38.75.75.241', '38.75.75.246', '38.75.75.251', '38.75.75.26', '38.75.75.29', '38.75.75.4',
              '38.75.75.44', '38.75.75.49', '38.75.75.56', '38.75.75.58', '38.75.75.62', '38.75.75.72',
              '38.75.75.76', '38.75.75.79', '38.75.75.88', '38.96.156.108', '38.96.156.112', '38.96.156.128',
              '38.96.156.131', '38.96.156.14', '38.96.156.142', '38.96.156.143', '38.96.156.149',
              '38.96.156.16',
              '38.96.156.163', '38.96.156.165', '38.96.156.169', '38.96.156.186', '38.96.156.188',
              '38.96.156.190',
              '38.96.156.192', '38.96.156.194', '38.96.156.199', '38.96.156.218', '38.96.156.236',
              '38.96.156.240',
              '38.96.156.252', '38.96.156.28', '38.96.156.32', '38.96.156.35', '38.96.156.56', '38.96.156.57',
              '38.96.156.6', '38.96.156.67', '38.96.156.77', '38.96.156.80', '38.96.156.83', '38.96.156.84',
              '38.96.156.89', '38.96.156.92', '45.238.157.100', '45.238.157.104', '45.238.157.106',
              '45.238.157.110', '45.238.157.116', '45.238.157.118', '45.238.157.119', '45.238.157.12',
              '45.238.157.123', '45.238.157.132', '45.238.157.14', '45.238.157.149', '45.238.157.15',
              '45.238.157.183', '45.238.157.186', '45.238.157.189', '45.238.157.2', '45.238.157.212',
              '45.238.157.214', '45.238.157.217', '45.238.157.22', '45.238.157.228', '45.238.157.23',
              '45.238.157.247', '45.238.157.43', '45.238.157.48', '45.238.157.51', '45.238.157.52',
              '45.238.157.53',
              '45.238.157.56', '45.238.157.61', '45.238.157.65', '45.238.157.72', '45.238.157.79',
              '45.238.157.8',
              '45.238.159.103', '45.238.159.107', '45.238.159.110', '45.238.159.114', '45.238.159.116',
              '45.238.159.123', '45.238.159.126', '45.238.159.144', '45.238.159.148', '45.238.159.15',
              '45.238.159.156', '45.238.159.165', '45.238.159.167', '45.238.159.183', '45.238.159.20',
              '45.238.159.208', '45.238.159.217', '45.238.159.220', '45.238.159.23', '45.238.159.230',
              '45.238.159.235', '45.238.159.237', '45.238.159.238', '45.238.159.24', '45.238.159.249',
              '45.238.159.251', '45.238.159.32', '45.238.159.34', '45.238.159.51', '45.238.159.6',
              '45.238.159.66',
              '45.238.159.77', '45.238.159.79', '45.238.159.82', '45.238.159.91']
        p_auth = str("csimonra:h19VA2xZ")
        p_host = random.choice(p1)
        p_port = "29842"
        proxy1 = {
            'http': "https://{}@{}:{}/".format(p_auth, p_host, p_port),
            'https': "http://{}@{}:{}/".format(p_auth, p_host, p_port)
        }

        response = requests.get(url=product_url, headers=headers,proxies=proxy1, timeout=30)
        # print(response.status_code)
        # exit()
        cache_data = response.text
        html_file_path = html_cache_page_saving(sku=str(prod_id_data[1]), saving_data=cache_data)
        import datetime
        time = datetime.datetime.now()
        dt_string1 =time.strftime("%m-%d-%YT%H:%M:%SZ")
        t1=str(dt_string1).replace('-13-','-12-')
        t2=str(dt_string1).replace('-13-','-11-')
        details_dict={'SKU_ID':prod_id_data[1],
                      'Website':'FOOTLOCKER',
                      'Country':'DE',
                      'RPC':'Not Available',
                      'MPC':'Not Available',
                      'Product_ID':'',
                      'Product_URL':product_url,
                      'Product_Name':'',
                      'Category_Path':'',
                      'Specification':'',
                      'Description':'',
                      'Currency':'EURO',
                      'List_Price':'',
                      'Promo_Price':'',
                      'Discount':'Not Available',
                      'Brand':'',
                      'Rating_Count':'Not Available',
                      'Review_Count':'Not Available',
                      'Image_URLs':'',
                      'Variant':'',
                      'Variant_ID':'-',
                      'Colour_of_Variant':'',
                      'Colour_Grouping':'',
                      'Seller_Name':'Not Available',
                      'Stock_Count':'Not Available',
                      'Stock_Condition':'',
                      'Stock_Message':'Not Available',
                      'Sustainability_Badge':'',
                      'Reason_Code':'',
                      'Crawling_TimeStamp':time.strftime("%m-%d-%YT%H:%M:%SZ"),
                      'Cache_Page_Link': html_file_path,
                      'Extra1': '-',
                      'Extra2': '-',
                      'Extra3': '-',
                      'Extra4': '-',
                      'Extra5': '-'
                      }

        if response.status_code == 200:
            try:
                tree=html.fromstring(response.content)
                all_json_data=re.findall('window.footlocker.STATE_FROM_SERVER =(.*);',response.text)
                # print(all_json_data)
                json_loads=json.loads(all_json_data[0])


                # cache_data = response.text
                # html_file_path = html_cache_page_saving(sku=str(prod_id_data[1]), saving_data=cache_data)

                #print(json_loads)
                #exit()
                prod_url_as_key=[ key for key in json_loads['details']['data']][0]
                print(prod_url_as_key)
                details_dict['Product_URL']='https://www.footlocker.de'+prod_url_as_key
                print('tttttttttttt',prod_url_as_key)

                if len(prod_url_as_key)==0:
                    details_dict['Product_URL'] = '-'
                #exit()
                # print(produrl)
                # details_dict['Category_Path']=
                image_sku=json_loads['details']['data'][prod_url_as_key][0]['imageSku'] # this gives the product id of the product
                details_dict['Product_ID']=image_sku
                if len(image_sku)==0:
                    details_dict['Product_ID'] = '-'
                details_dict['Colour_of_Variant']=json_loads['details']['data'][prod_url_as_key][0]['style'].replace('|','')
                if len(str(details_dict['Colour_of_Variant']))==0:
                    details_dict['Colour_of_Variant'] = '-'
                prod_name=json_loads['details']['product'][prod_url_as_key]['name']
                x = prod_name.split(' - ')
                x=x[0]
                #print(x)
                #exit()

                #print(prod_name)
                details_dict['Product_Name']=x #prod_name
                if len(details_dict['Product_Name'])==0:
                    details_dict['Product_Name']='-'
                details_dict['Product_Name'] = general.clean(details_dict['Product_Name'])
                brand=json_loads['details']['product'][prod_url_as_key]['brand']
                #print(brand)
                details_dict['Brand']=brand
                if len(brand)==0:
                    details_dict['Brand']='-'
                details_dict['Brand']= general.clean(details_dict['Brand'])

                l_price=json_loads['details']['data'][prod_url_as_key][0]['price']['originalPrice']
                details_dict['List_Price']=l_price
                if len(str(l_price))==0:
                    details_dict['List_Price']='-'


                # l_price = tree.xpath('//span[@class="ProductPrice-original"]/text()')
                # details_dict['List_Price'] = l_price
                promo_price_ = tree.xpath('//span[@class="ProductPrice-final"]/text()')
                p_price = json_loads['details']['data'][prod_url_as_key][0]['price']['value']
                # p_price = tree.xpath('//span[@class="ProductPrice-final"]/text()')
                details_dict['Promo_Price']=p_price
                if len(str(p_price))==0:
                    details_dict['Promo_Price']='-'


                all_img=[]
                for img_no in range(1,8):
                    image='https://images.footlocker.com/is/image/FLEU/{}_0{}?wid=503&hei=503&fmt=png-alpha'.format(image_sku,img_no)
                    all_img.append(image)
                #print(all_img)
                details_dict['Image_URLs']='|'.join(i for i in all_img)
                if len(details_dict['Image_URLs'])==0:
                    details_dict['Image_URLs']='-'


                col_grp=[]
                for data in json_loads['details']['styles'][prod_url_as_key]:
                    col_grp.append(data['name'])
                details_dict['Colour_Grouping']='|'.join(i.replace('|','') for i in col_grp)
                if len(details_dict['Colour_Grouping']) == 0:
                    details_dict['Colour_Grouping'] = '-'
                details_dict['Colour_Grouping'] = general.clean(details_dict['Colour_Grouping'])
                #description

                # disc=json_loads['details']['product'][prod_url_as_key]['description']
                #
                # disc_val= re.findall('<p>(.*?)</p>',disc)
                # if len(disc_val)>0:
                #     details_dict['Description'] =disc_val[1]
                # else:
                #     details_dict['Description'] ='-'
                #
                # new=re.findall('<li>&nbsp;(.*?)</li>',disc)
                # details_dict['Specification']='|'.join(i for i in new)
                disc=tree.xpath('//div[@class="ProductDetails-description"]/p/text()')
                #print(disc)

                details_dict['Description']=''.join(val for val in disc if '\xa0' not in val)
                if len(details_dict['Description'])==0:
                    details_dict['Description']='-'
                details_dict['Description']=general.clean(details_dict['Description'])

                spec=tree.xpath('//div[@class="ProductDetails-description"]/ul/li/text()')
                details_dict['Specification']='|'.join(i.replace('\xa0','') for i in spec)
                if len(details_dict['Specification'])==0:
                    details_dict['Specification']='-'
                details_dict['Specification']=general.clean(details_dict['Specification'])

                details_dict['Reason_Code']='Success-PF'

                sust_badge=tree.xpath('//div[@class="Product-header"]/div/div/img/@src')

                #print(sust_badge)
                new=tree.xpath('//div[@class="Product-header"]/div')
                #print(new)

                if 'Sustainability' in response.text:
                    details_dict['Sustainability_Badge']='Yes'
                else:
                    details_dict['Sustainability_Badge']='No'

                # print(new)
                cat=tree.xpath('//nav[@class="c-breadcrumbs"]/ol/li/text()')
                #print(cat)
                details_dict['Category_Path']='Home >'+'>'.join(i for i in cat)
                if len(details_dict['Category_Path'])==0:
                    details_dict['Category_Path']='-'
                details_dict['Category_Path'] = general.clean(details_dict['Category_Path'])

                size_info_html=tree.xpath('//div[@class="ProductSize-group"]/div')
                # print(size_info_html)
                # print(len(size_info_html))
                for element in size_info_html:
                    ob1=element.xpath('.//label/span/text()')[0]
                    details_dict['Variant']=ob1
                    details_dict['Variant'] = general.clean(details_dict['Variant'])
                    ob2 = element.xpath('.//input/@aria-label')  #in this also contain product size if

                    if 'out of stock' in ob2[0]:
                        details_dict['Stock_Condition']="Out of Stock"
                    else:
                        details_dict['Stock_Condition']='In Stock'

                    print(details_dict)
                    store_data_by_product_id([details_dict])

                    # print(ob1,ob2)
            except Exception as e:
                details_dict = {'SKU_ID': prod_id_data[1],
                                'Website': 'FOOTLOCKER',
                                'Country': 'DE',
                                'RPC': 'Not Available',
                                'MPC': 'Not Available',
                                'Product_ID': '-',
                                'Product_URL': '-',
                                'Product_Name': '-',
                                'Category_Path': '-',
                                'Specification': '-',
                                'Description': '-',
                                'Currency': 'EURO',
                                'List_Price': '-',
                                'Promo_Price': '-',
                                'Discount': 'Not Available',
                                'Brand': '-',
                                'Rating_Count': 'Not Available',
                                'Review_Count': 'Not Available',
                                'Image_URLs': '-',
                                'Variant': '-',
                                'Variant_ID': '-',
                                'Colour_of_Variant': '-',
                                'Colour_Grouping': '-',
                                'Seller_Name': 'Not Available',
                                'Stock_Count': 'Not Available',
                                'Stock_Condition': '-',
                                'Stock_Message': 'Not Available',
                                'Sustainability_Badge': '-',
                                'Reason_Code': 'Success-PNF',
                                'Crawling_TimeStamp': time.strftime("%Y-%m-%dT%H:%M:%SZ"),
                                'Cache_Page_Link': html_file_path,
                                'Extra1': '-',
                                'Extra2': '-',
                                'Extra3': '-',
                                'Extra4': '-',
                                'Extra5': '-'
                                }
                store_data_by_product_id([details_dict])
        elif n != 0:
            new_n = n - 1
            by_product_id(prod_id_data,n=new_n)
            store_data_by_product_id([details_dict])
        elif response.status_code == 408:
            print('TimeOut')
            store_data_by_product_id([details_dict])
        else:
            print('Blocked')
            store_data_by_product_id([details_dict])

    except:
        details_dict['Reason_Code'] = 'Success-PNF'
        store_data_by_product_id([details_dict])
        print('PNF')







for data in extract_xl():
    print('inside 1')
    by_product_id(prod_id_data=data,n=3)
    # break






